from flask import Flask, request, send_file
from openpyxl import load_workbook
import requests
import io
import os

app = Flask(__name__)

@app.route('/fill-amtrust', methods=['POST'])
def fill_amtrust():
    data = request.json
    
    # Download the template from Supabase storage
    template_url = data['template_url']
    template_response = requests.get(template_url)
    template_bytes = io.BytesIO(template_response.content)
    
    wb = load_workbook(template_bytes)
    ws = wb["HOA App"]
    
    # Header fields
    ws["C4"] = data.get("named_insured", "")
    ws["C5"] = data.get("fein", "")
    ws["C6"] = data.get("property_mgmt", "")
    ws["C8"] = data.get("contact_name", "")
    ws["C9"] = data.get("contact_phone", "")
    ws["C11"] = data.get("hoa_address", "")
    ws["C12"] = data.get("homes_completed", 0)
    ws["C13"] = data.get("pools", 0)
    ws["C14"] = data.get("ponds", 0)
    ws["C15"] = data.get("miles_road", 0)
    
    # Category rollups into row 22
    ws["A22"] = data.get("fences_monuments", 0)
    ws["B22"] = data.get("parks_recreation", 0)
    ws["C22"] = data.get("pools_equipment", 0)
    ws["D22"] = data.get("street_lighting", 0)
    ws["E22"] = data.get("clubhouse", 0)
    
    # Other items
    other_items = data.get("other_items", [])
    for i, item in enumerate(other_items):
        row = 22 + i
        ws[f"G{row}"] = item.get("value", 0)
        ws[f"H{row}"] = item.get("name", "")
    
    # Force recalculation on open
    wb.calculation.fullCalcOnLoad = True
    
    # Save to buffer and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{data.get('named_insured', 'export')}_AmTrust.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
